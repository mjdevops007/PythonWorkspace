import openpyxl

f = openpyxl.load_workbook("inventory.xlsx")

inv_list = f["Sheet1"]

print(inv_list.max_row)

# print number of products per supplier

supplier_product_list = {}
supplier_inventory_list = {}
product_list = []

for i in range(2, inv_list.max_row + 1):
    supplier_name = inv_list.cell(i, 4).value
    inventory_value = inv_list.cell(i, 2).value
    price = inv_list.cell(i, 3).value

    # add the additional column to capture total inventory price for each product
    inv_list.cell(i,5).value = inventory_value * price

    if supplier_name in supplier_product_list:
        supplier_product_list[supplier_name] = supplier_product_list[supplier_name] + 1
    else:
        print("adding new supplier")
        supplier_product_list[supplier_name] = 1

    # print total inventory cost per supplier
    if supplier_name in supplier_inventory_list:
        supplier_inventory_list[supplier_name] = supplier_inventory_list[supplier_name] + (inventory_value * price)
    else:
        print("adding new supplier")
        supplier_inventory_list[supplier_name] = inventory_value * price

    # print product list with price less than 10

    if price < 10:
        product_list.append(i)


inv_list.cell(1,5).value = "Total_Value"
print(f"Supplier list with product count: {supplier_product_list}")
print(f"Supplier list with total inventory price: {supplier_inventory_list}")
print(f"Product list with price less than 10 {product_list}")

#save the workbook
f.save("inventory_with_total_value.xlsx")